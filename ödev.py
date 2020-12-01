# -*- coding: utf-8 -*-
import requests as req
import xlsxwriter
from bs4 import BeautifulSoup
import numpy as np
import pandas as pd
from sklearn.impute import SimpleImputer
from sklearn.model_selection import train_test_split
from sklearn import preprocessing
from openpyxl import load_workbook as lw

# kullanılacak veri yapıları
brandList = list()
yearList = list()
kmList = list()
priceList = list()

# bilgilerin çekilmesi
res = req.get('https://www.arabam.com/ikinci-el/otomobil?take=50&page=10')

pack = res.content

soup = BeautifulSoup(pack, "html.parser")

# çekilen bilgileri tag şeklinden stringe çeviriyoruz
brands = soup.select('.listing-list-item td:nth-child(2)')

for brand in brands:
    brandList.append(brand.text.split(' ', 1)[0])
    
years = soup.select('.listing-list-item td:nth-child(4)')

for year in years:
    yearList.append(str(year.text))
    
kmS = soup.select('.listing-list-item td:nth-child(5)')

for km in kmS:
    kmList.append(str(km.text))
    
prices = soup.select('.listing-list-item td:nth-child(7)')

for price in prices:
    priceList.append(str(price.text.replace(' TL', '')))
    
# verileri çekip düzenledikten sonra excele yazdırma işlemleri
# dosyayı oluşturduk
outWorkbook = xlsxwriter.Workbook("arabaData.xlsx")
outSheet = outWorkbook.add_worksheet()

# verileri excele yazdırdık
outSheet.write("A1", "Brand")
outSheet.write("B1", "Year")
outSheet.write("C1", "KM")
outSheet.write("D1", "Price")

for counter in range(len(brandList)):
    outSheet.write(counter+1, 0, brandList[counter])
    outSheet.write(counter+1, 1, yearList[counter])
    outSheet.write(counter+1, 2, kmList[counter])
    outSheet.write(counter+1, 3, priceList[counter])
    
outWorkbook.close()

# burda xlsx dosyasını csv ye döndürdükten sonra gelen çift tırnakları sildik
with open('odev_araba.csv', "r+", encoding="utf-8") as csv_file:
    content = csv_file.read()

with open('odev_araba.csv', "w+", encoding="utf-8") as csv_file:
    csv_file.write(content.replace('"', ''))
    
# ön işleme kısmı
# verileri çekiyoruz ve one hot encoding yapıyoruz
dataset = pd.read_csv('asil_veri.csv')
rated_dummies = pd.get_dummies(dataset.Brand)
dataset = pd.concat([rated_dummies, dataset], axis=1)
dataset.drop('Brand',inplace=True,axis=1)
dep = dataset.iloc[:, :-1].values
indep = dataset.iloc[:, -1].values

# eksik olan kısımları dolduruyoruz 
imputer = SimpleImputer(missing_values=np.nan, strategy='mean')
imputer.fit(dep[:, -2:])
dep[:, -2:] = imputer.transform(dep[:, -2:])

# train test kısmı
dep_train, dep_test, indep_train, indep_test = train_test_split(dep, indep, test_size = 0.2, random_state = 1)

# scale etme kısmı, sadece km'yi scale ediyoruz

# train km scale etme
min_max_scaler = preprocessing.MinMaxScaler()
train_km = dep_train[:, -1].reshape(-1,1)
train_km = min_max_scaler.fit_transform(train_km)
dep_train[:, -1] = train_km.ravel()

# test km scale etme
test_km = dep_test[:, -1].reshape(-1,1)
test_km = min_max_scaler.fit_transform(test_km)
dep_test[:, -1] = test_km.ravel()

# işlemleri tamamlanan veriler excele yazdırılıyor    
wb = lw('odev_araba.xlsx')
sheets = wb.sheetnames
af_pre = wb[sheets[1]]
for ctr in range(500):
   for ctr2 in range(27):
       af_pre.cell(ctr+2, ctr2+1).value = dep[ctr,ctr2]
       
list_test_km = test_km.tolist()
list_train_km = train_km.tolist()

sayac_km = 0
sayac_price = 0

for km in list_test_km:
     af_pre.cell(sayac_km+2, 28).value = str(km)    
     sayac_km += 1

for km in list_train_km:
    af_pre.cell(sayac_km+1, 28).value = str(km)
    sayac_km += 1

for price in indep:
    af_pre.cell(sayac_price+2, 29).value = price
    sayac_price += 1

wb.save('odev_araba.xlsx')
wb.close()

############## SON #####################






    
    






